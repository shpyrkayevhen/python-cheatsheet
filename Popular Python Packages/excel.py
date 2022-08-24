# 1: pip install openpyxl
import openpyxl

# Have one empty sheet obj
# wb = openpyxl.Workbook()

wb = openpyxl.load_workbook('transactions.xlsx')
# print(wb.sheetnames)

sheet = wb["Sheet1"]


# Method sheet obj

# (1) Add a row with the values into the end our sheet
# sheet.append(["1value", "2value", "3value"])


# (2) Add the row|column into the neсessary place on index
# sheet.insert_rows(["1value", "2value", "3value"], 3)
# sheet.insert_cols()


# (3) Delete the row|column
# sheet.delete_rows(2)
# sheet.delete_cols(1)


# (4) Save wb as a new file
wb.save("transactions2.xlsx")


# Path a row and column
# The best way щоб динамічно отримувати доступ
# cell = sheet.cell(row=1, column=1)


cell = sheet["a1"]


# (3) Return all ceils in this column
column = sheet["a"]
print(column)

# for ceil in column:
#    print(ceil.value)


# (4) Return all ceils in the range this columns
cells = sheet["a:c"]

for column in cells:
    for ceil in column:
        print(ceil.value)


# (5) Ruturn all cells in this cells range
sheet["a1:c3"]

# (6) Returns all cells in 1 rows OR 1-3
row = sheet[1]
print(row)

ceils = sheet[1:3]


# sheet.max_row
# sheet.max_column

# for row in range(1, sheet.max_row + 1):
#    for column in range(1, sheet.max_column + 1):
#       cell = sheet.cell(row, column)
#       print(cell.value)

# cell - клітинка
# cell = sheet["a1"]
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
# We can change a value
# cell.value = 'id'


# Creat new sheet in wb and put it before 'Sheet1'
# wb.create_sheet('Sheet2', 0)
# sheet2 = wb['Sheet2']


# Remove the existing sheet
# wb.remove_sheet(sheet2)


# Get all vallue(cells) in 1 colbmn
# for row in range(1, 5):
#     cell = sheet.cell(row, 1)
#     print(cell.value)
